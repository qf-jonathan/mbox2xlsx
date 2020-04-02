import mailbox
from email.header import decode_header
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from mimetypes import guess_extension
import pathlib
import uuid
import re
import ntpath
from datetime import datetime


def get_subject(message):
    if message['subject'] is not None:
        subject, encoding = decode_header(message['subject'])[0]
        if encoding is not None:
            subject = subject.decode(encoding)
        if subject.startswith('"') and subject.endswith('"'):
            subject = subject.strip('"')
        return subject
    return ''


def get_email_header(message, key, email_only=False):
    header_value = message[key]
    if header_value is None:
        return ''
    header, encode = decode_header(header_value)[0]
    if encode:
        header = header.decode(encode)
    if email_only:
        email_list = re.findall(r'[\w\s]*<([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)>', str(header))
        if len(email_list) > 0:
            return email_list[0]
        email_list = re.findall(r'[\w\s]*<([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)>', header_value)
        if len(email_list) > 0:
            return email_list[0]
        return header_value
    return header


def get_payload(message, prefix='', message_id='', first_only=False, hyperlinked=True):
    if not message.is_multipart():
        filename = message.get_filename()
        extension = guess_extension(message.get_content_type(), False)
        file_id = f'{message_id}{str(uuid.uuid1())[0:8]}'
        if extension is None or extension == '.bat':
            extension = '.txt'
        if filename is None:
            filename = f'{file_id}{extension}'
        else:
            filename, encode = decode_header(filename)[0]
            if encode is not None:
                filename = filename.decode(encode)
            filename = f'{file_id}-{filename}'
        file_path = f'{prefix.rstrip("/")}/{filename}'
        with open(file_path, 'wb') as file:
            file.write(message.get_payload(decode=True))
        if hyperlinked:
            return [f'=HYPERLINK("{file_path}", "{filename}")']
        return [filename]
    ret = []
    for message_part in message.get_payload():
        part = get_payload(message_part, prefix, message_id, first_only, hyperlinked)
        if first_only:
            return part
        ret += part
    return ret


def get_payload_data(message):
    if not message.is_multipart():
        if message.get_content_type() != 'text/plain':
            return [None, None, None]
        content = message.get_payload(decode=True)
        charset = message.get_param('charset')
        try:
            if charset is not None:
                content = content.decode(charset)
        except:
            pass
        contact = re.findall(r'[Dd]e\s*:([\w\s]*)<([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)>\s*\n', str(content))
        whatsapp = re.findall(r'[Ww][Hh][Aa][Tt][Ss][Aa][Pp][Pp]\s*:\s*(\+?\d[\d-]{5,20}\d)', str(content))
        if len(contact) > 0:
            ret = [contact[0][0].strip(), contact[0][1], None]
            if len(whatsapp) > 0:
                ret[2] = whatsapp[0]
            return ret
        return [None, None, None]
    for message_part in message.get_payload():
        data = get_payload_data(message_part)
        if data[0] is not None:
            return data
    return [None, None, None]


def proccess_mbox(file_path, date, ito, eto, ifrom, efrom, cc, subject, body, contact_data, attachment, tick_fn=None):
    base_dir, filename = ntpath.split(file_path)
    filename_base = "".join(filename.split(".")[0:-1])
    end_path = f'{base_dir}/{filename_base}({datetime.now()})'
    mbox = mailbox.mbox(file_path)
    wb = Workbook()
    ws = wb.active
    curr, total = 0, len(mbox)
    len_id = len(str(total))
    ws.append([])
    excel_header, excel_header_width, first_row = [], [], True
    mx_attachment = 0
    for message in mbox:
        row = []
        if date:
            if first_row:
                excel_header.append('Fecha')
                excel_header_width.append(8)
            row.append(message['date'])
        if ito:
            if first_row:
                excel_header.append('Destinatario')
                excel_header_width.append(15)
            row.append(get_email_header(message, 'to', email_only=eto))
        if ifrom:
            if first_row:
                excel_header.append('Remitente')
                excel_header_width.append(15)
            row.append(get_email_header(message, 'from', email_only=efrom))
        if cc:
            if first_row:
                excel_header.append('Con Copia')
                excel_header_width.append(15)
            row.append(get_email_header(message, 'cc'))
        if subject:
            if first_row:
                excel_header.append('Asunto')
                excel_header_width.append(20)
            row.append(get_subject(message))
        if contact_data:
            if first_row:
                excel_header.append('Nombre')
                excel_header.append('Correo')
                excel_header.append('Whatsapp')
                excel_header_width += [15, 15, 15]
            row += get_payload_data(message)
        if body and not attachment:
            if first_row:
                excel_header.append('Contenido')
                excel_header_width.append(20)
            body_path = f'{end_path}/adjunto'
            pathlib.Path(body_path).mkdir(exist_ok=True, parents=True)
            row += get_payload(message, message_id=f'{curr:0{len_id}}', prefix=body_path, first_only=True)
        elif attachment:
            if first_row:
                excel_header.append('Contenido')
                excel_header_width.append(20)
            body_path = f'{end_path}/adjunto'
            pathlib.Path(body_path).mkdir(exist_ok=True, parents=True)
            attachment_list = get_payload(message, message_id=f'{curr:0{len_id}}', prefix=body_path)[1:]
            row += attachment_list
            mx_attachment = max(mx_attachment, len(attachment_list) - 1)
        first_row = False
        ws.append(row)
        curr += 1
        if tick_fn:
            tick_fn(curr, total)
    for i in range(mx_attachment):
        excel_header.append(f'Adjunto {i + 1}')
        excel_header_width.append(15)
    for i in range(len(excel_header)):
        ws.cell(1, i + 1, excel_header[i]).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(i + 1)].width = excel_header_width[i]
    pathlib.Path(end_path).mkdir(exist_ok=True)
    wb.save(f'{end_path}/{filename_base}.xlsx')
